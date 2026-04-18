from __future__ import annotations

from pathlib import Path

from models import ExtractedContent, ScanConfig

from .binary_extractor import extract_binary_strings
from .common import finalize_text, safe_import


XLRD = safe_import("xlrd")
OPENPYXL = safe_import("openpyxl")


def extract_xls(path: Path, config: ScanConfig) -> ExtractedContent:
    warnings: list[str] = []
    suffix = path.suffix.lower()

    if suffix == ".xlsx" and OPENPYXL is not None:
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
            rows: list[str] = []
            for sheet in workbook.worksheets:
                rows.append(f"[sheet] {sheet.title}")
                for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                    rows.append(" | ".join("" if value is None else str(value) for value in row))
                    if row_index + 1 >= config.max_structured_rows:
                        rows.append("[truncated]")
                        break
            return finalize_text("\n".join(rows), "xlsx_openpyxl", config.max_text_chars)
        except Exception:
            warnings.append("openpyxl не смог прочитать XLSX, использован fallback.")

    if XLRD is not None:
        try:
            import xlrd

            workbook = xlrd.open_workbook(path)
            rows = []
            for sheet in workbook.sheets():
                rows.append(f"[sheet] {sheet.name}")
                for row_index in range(min(sheet.nrows, config.max_structured_rows)):
                    rows.append(" | ".join(str(sheet.cell_value(row_index, col)) for col in range(sheet.ncols)))
            result = finalize_text("\n".join(rows), "xls_xlrd", config.max_text_chars)
            result.warnings.extend(warnings)
            return result
        except Exception:
            warnings.append("XLRD не смог прочитать XLS, использован binary fallback.")

    result = extract_binary_strings(path, config)
    result.method = "spreadsheet_binary_fallback"
    result.warnings.extend(warnings)
    return result
